import json
from flask import Blueprint, request, send_file
from website.helpers.require_role import require_role_system_name_on_current_user
from flask_login import current_user
from website.models.term import Term
from website.models.deck import Deck
from website.models.answer import Answer
import openpyxl


slovnik_api = Blueprint("slovnik_api", __name__)


@slovnik_api.route("/terms")
@require_role_system_name_on_current_user("user")
def get_terms():
    return json.dumps(Term.get_all_for_user_list())


@slovnik_api.route("/decks")
@require_role_system_name_on_current_user("user")
def get_decks():
    return json.dumps(Deck.get_all_for_user_list())


@slovnik_api.route("/deck/<int:deck_id>")
@require_role_system_name_on_current_user("user")
def get_terms_in_deck(deck_id):
    deck = Deck.get_by_id(deck_id)
    if deck.author_id != current_user.id:
        return json.dumps({"error": "Nemáte oprávnění zobrazit tento balíček!"})
    return json.dumps(deck.for_detail())


@slovnik_api.route("/quiz/<int:deck_id>")
@require_role_system_name_on_current_user("user")
def get_quiz(deck_id):
    deck = Deck.get_by_id(deck_id)
    if deck.author_id != current_user.id:
        return json.dumps({"error": "Nemáte oprávnění zobrazit tento balíček!"})
    return json.dumps(deck.for_quiz())

@slovnik_api.route("/export_all")
@require_role_system_name_on_current_user("user")
def export_all():
    # xlsx file with four sheets:
    # terms: id, datetime in isoformat, front, back, times_tested, times_correct
    # decks: id, datetime in isoformat, name
    # terms_decks: term_id, deck_id
    # answers: id, term_id, value
    answers = []
    
    workbook = openpyxl.Workbook()
    terms_sheet = workbook.active
    terms_sheet.title = "terms"
    terms_sheet.append(["id", "datetime", "front", "back", "times_tested", "times_correct"])
    for term in Term.get_all_by_user(current_user.id):
        answers.extend(term.answers)
        terms_sheet.append([
            term.id,
            term.datetime.isoformat(),
            term.front,
            term.back,
            term.times_tested,
            term.times_correct
        ])
    
    decks_sheet = workbook.create_sheet("decks")
    decks_sheet.append(["id", "datetime", "name"])
    for deck in Deck.get_all_by_user(current_user.id):
        decks_sheet.append([
            deck.id,
            deck.datetime.isoformat(),
            deck.name
        ])
    
    terms_decks_sheet = workbook.create_sheet("terms_decks")
    terms_decks_sheet.append(["term_id", "deck_id"])
    for deck in Deck.get_all_by_user(current_user.id):
        for term in deck.terms:
            terms_decks_sheet.append([term.id, deck.id])
    
    answers_sheet = workbook.create_sheet("answers")
    answers_sheet.append(["id", "term_id", "value"])
    for answer in answers:
        answers_sheet.append([
            answer.id,
            answer.term_id,
            answer.value
        ])
        
    from io import BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="slovnik_export.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@slovnik_api.route("/decks_select")
@require_role_system_name_on_current_user("user")
def get_decks_select():
    result = []
    for deck in Deck.get_all_by_user(current_user.id):
        result.append({
            "id": deck.id,
            "name": deck.name
        })
    return json.dumps(result)


@slovnik_api.route("/export_deck/<int:deck_id>")
@require_role_system_name_on_current_user("user")
def export_deck(deck_id):
    # xlsx file with one sheet:
    # no header, columns: front, back
    deck = Deck.get_by_id(deck_id)
    if deck.author_id != current_user.id:
        return json.dumps({"error": "Nemáte oprávnění zobrazit tento balíček!"})
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for term in deck.terms:
        sheet.append([term.front, term.back])
    from io import BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"{deck.name}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")